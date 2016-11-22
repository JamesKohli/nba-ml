import openpyxl
import csv
from datetime import datetime

data = openpyxl.load_workbook('data/Team-sample.xlsx', read_only=True)
ws = data.active

with open('data/teamdata.csv', 'wb') as csvfile:
    datawriter = csv.writer(csvfile, delimiter=',')
    datawriter.writerow(['reg_season', 'home_team', 'away_team', 'game_date', 'away_player_one', 'away_player_two',
                             'away_player_three', 'away_player_four', 'away_player_five', 'home_player_one', 'home_player_two',
                             'home_player_three', 'home_player_four', 'home_player_five', 'head_ref', 'away_rest_days',
                             'home_rest_days', 'spread','result'])

    rowcount = ws.max_row
    #for each row in ws
    for x in range(2, rowcount, 2):

        if 'Regular Season' in ws['A'+str(x)].value:
            reg_season = True
        else:
            reg_season = False
        away_team = ws['C'+str(x)].value
        home_team = ws['C'+str(x+1)].value
        game_date = datetime.strptime(ws['B' + str(x)].value,'%m/%d/%Y')
        away_player_one = ws['AJ' + str(x)].value.replace("'","")
        away_player_two = ws['AK' + str(x)].value.replace("'","")
        away_player_three = ws['AL' + str(x)].value.replace("'","")
        away_player_four = ws['AM' + str(x)].value.replace("'","")
        away_player_five = ws['AN' + str(x)].value.replace("'","")
        home_player_one = ws['AJ' + str(x+1)].value.replace("'","")
        home_player_two = ws['AK' + str(x+1)].value.replace("'","")
        home_player_three = ws['AL' + str(x+1)].value.replace("'","")
        home_player_four = ws['AM' + str(x+1)].value.replace("'","")
        home_player_five = ws['AN' + str(x+1)].value.replace("'","")
        head_ref = ws['AO' + str(x)].value
        if ws['AI' + str(x)].value == '3+':
            away_rest_days = 3
        else:
            away_rest_days = ws['AI' + str(x)].value
        if ws['AI' + str(x+1)].value == '3+':
            home_rest_days = 3
        else:
            home_rest_days = ws['AI' + str(x+1)].value
        spread = ws['AR'+str(x)].value
        if ws['M'+str(x+1)].value - ws['M'+str(x)].value < spread:
            away_cover = True
        else:
            away_cover = False
        if ws['M' + str(x)].value - ws['M' + str(x+1)].value > spread * -1:
            result = "away_cover"
        elif ws['M' + str(x+1)].value - ws['M' + str(x)].value > spread:
            result = "home_cover"
        else:
            result = "push"

        datawriter.writerow([reg_season, home_team, away_team, game_date, away_player_one, away_player_two,
                             away_player_three, away_player_four, away_player_five, home_player_one, home_player_two,
                             home_player_three, home_player_four, home_player_five, head_ref, away_rest_days,
                             home_rest_days,spread,result])
        #datawriter.writerow([Home Team, Away Team, Date, Home 1, Home 2, Home 3, Home 4, Home 5, Away 1, Away 2, Away 3, Away 4, Away 5, Head Ref, Regular Season, Rest Days, Line, Home Wins, Away Wins, Push])
